from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl

driver = webdriver.Chrome('./chromedriver')
driver.maximize_window()
driver.get("https://promo.blackanddecker.ae/nasrin/")

def automation_process():
    print("Operation Started")
    excel_open()
    print("Operation Finished")
    driver.close()
    

def excel_open():
    path = "G:\Automation\Automation\\Test.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row=sheet_obj.max_row
    max_col = sheet_obj.max_column
    
    for i in range(2,max_row+1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)
            print(cell_obj.coordinate,cell_obj.value)
            if "A" in cell_obj.coordinate:
                name = cell_obj.value
            elif "B" in cell_obj.coordinate:
                lastname = cell_obj.value
            else:
                email = cell_obj.value
        print("Iterating for row"+str(i)+";column"+str(j))
        form_fill(name,lastname,email)  
        

def form_fill(name,lastname,email):
    driver.find_element(By.ID, "Name").click()
    driver.find_element(By.ID, "Name").send_keys(name)
    driver.find_element(By.ID, "Last_Name").click()
    driver.find_element(By.ID, "Last_Name").send_keys(lastname)
    driver.find_element(By.ID, "email").click()
    driver.find_element(By.ID, "email").send_keys(email)
    driver.find_element(By.ID, "Hobbies_or_Interests_house_cleaning").click()
    driver.find_element(By.ID, "Hobbies_or_Interests_gardening").click()
    driver.find_element(By.ID, "Hobbies_or_Interests_home_improvement").click()
    driver.find_element(By.ID, "Hobbies_or_Interests_diy").click()
    driver.find_element(By.ID, "Hobbies_or_Interests_cooking").click()
    driver.find_element(By.ID, "Opt_In_Marketing_yes").click()
    driver.find_element(By.CSS_SELECTOR, ".label > strong").click()
    try:
        element_present = EC.presence_of_element_located((By.CSS_SELECTOR, '.close__2NcGKV'))
        WebDriverWait(driver, 3).until(element_present)
        driver.find_element(By.CSS_SELECTOR, ".close__2NcGKV").click()
    except TimeoutException:
        print("Timed out waiting for page to load")
    
    

automation_process()